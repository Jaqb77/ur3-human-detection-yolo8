import time
import os
import queue
import threading
from datetime import datetime
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import cv2  
import pickle  
import config

class excel_logger:
    
    def __init__(self):
        self.headers = [
            "id_detection",
            "id_camera",
            "status",
            "violated_zone",
            "ai_model",
            "hardware_setup",
            "detection_start_time",
            "detection_end_time",
            "duration_s",
            "average_fps",
            "average_model_accuracy",
            "average_inference_time_ms",
            "tcp_x",
            "tcp_y",
            "tcp_z"
        ]

        # Bufor w pamięci głównej oraz kolejka asynchroniczna do zapisu w wątku roboczym
        self.log_queue = queue.Queue()
        self.is_active = False          
        self.start_time = 0.0           
        self.start_timestamp_str = ""  
        self.next_id = 1
        self.current_det_id = 1
        
        self.session_accuracies = []    
        self.session_fps = []           
        self.session_inference_times = [] 
        
        self.zone_priority = {"NONE": 0, "GREEN": 1, "YELLOW": 2, "RED": 3}
        self.highest_zone = "NONE"
        self.detection_frame = None
        
        self.db_dir = os.path.dirname(os.path.abspath(__file__))
        self.log_file_path = ""
        
        # Uruchomienie wątku roboczego do asynchronicznego zapisu plików Excel i obrazów
        self.worker_thread = threading.Thread(target=self._async_writer, daemon=True)
        self.worker_running = True
        self.worker_thread.start()

    def _get_absolute_path(self, filename):
        if not os.path.isabs(filename) and os.path.basename(filename) == filename:
            return os.path.join(self.db_dir, filename)
        return filename

    def init_check(self, filename):
        self.log_file_path = self._get_absolute_path(filename)
        cache_path = self.log_file_path + ".cache"
        
        # 1. Odzyskiwanie z pliku cache, jeśli istnieje
        if os.path.exists(cache_path):
            try:
                with open(cache_path, 'rb') as f:
                    cached_data = pickle.load(f)
                
                if os.path.exists(self.log_file_path):
                    wb = load_workbook(self.log_file_path)
                    ws = wb.active
                else:
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "datalogs"
                    ws.append(self.headers)
                    self._apply_formatting(ws)

                for row in cached_data.get('buffer', []):
                    ws.append(row)
                    
                self._apply_formatting(ws)
                wb.save(self.log_file_path)
                
                # Odzyskiwanie zapisanych klatek
                image_buf = cached_data.get('image_buffer', {})
                if image_buf:
                    base_dir = os.path.dirname(self.log_file_path)
                    folder_name, _ = os.path.splitext(os.path.basename(self.log_file_path))
                    target_folder = os.path.join(base_dir, folder_name)
                    os.makedirs(target_folder, exist_ok=True)
                    for det_id, (frame, safe_timestamp) in image_buf.items():
                        img_filename = f"{det_id} {safe_timestamp}.jpg"
                        img_path = os.path.join(target_folder, img_filename)
                        cv2.imwrite(img_path, frame)
                
                os.remove(cache_path)
                print("[Excel] Pomyślnie przywrócono zaległe dane z pliku cache.")
            except PermissionError:
                print("\n[Excel][WARNING] Główny plik Excel jest otwarty. Dane pozostają w cache.\n")
            except Exception as e:
                print(f"[Excel][ERROR] Nie udało się odzyskać danych z cache: {e}")

        # 2. Ustalenie kolejnego ID na podstawie pliku Excel
        if os.path.exists(self.log_file_path):
            try:
                wb = load_workbook(self.log_file_path)
                ws = wb.active
                if ws.max_row > 1:
                    last_id = ws.cell(row=ws.max_row, column=1).value
                    self.next_id = int(last_id) + 1
                else:
                    self.next_id = 1
            except Exception as e:
                print(f"[Excel][ERROR] Nie można odczytać pliku: {e}")
                self.next_id = 1
        else:
            try:
                wb = Workbook() 
                ws = wb.active
                ws.title = "datalogs"
                ws.append(self.headers)
                self._apply_formatting(ws)
                wb.save(self.log_file_path)
                self.next_id = 1
            except Exception as e:
                print(f"[Excel][ERROR] Błąd przy tworzeniu nowego pliku Excel: {e}")

    def acquisition(self, detection_bool, camera_id, current_acc, model_name, hardware_setup, current_fps, current_inf_time, robot_x, robot_y, robot_z, active_zone_name="NONE", frame=None, det_id=1):
        """
        Główna metoda zbierania danych wywoływana w każdej klatce.
        """
        if detection_bool and not self.is_active:
            self.is_active = True
            self.current_det_id = det_id
            self.start_timestamp_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self.start_time = time.time()
            self.highest_zone = active_zone_name
            if frame is not None:
                self.detection_frame = frame.copy() 
                # Natychmiast wysyłamy pierwsze zdjęcie do zapisu w czasie rzeczywistym
                safe_timestamp = self.start_timestamp_str.replace(":", "-")
                self.log_queue.put({
                    "type": "save_image",
                    "frame": self.detection_frame,
                    "safe_timestamp": safe_timestamp,
                    "det_id": det_id
                })
            self.session_accuracies.append(current_acc)
            self.session_fps.append(current_fps)
            self.session_inference_times.append(current_inf_time)

        elif detection_bool and self.is_active:
            # Jeśli naruszono strefę o wyższym priorytecie, aktualizujemy zdjęcie na to z najgłębszego wejścia
            if self.zone_priority.get(active_zone_name, 0) > self.zone_priority.get(self.highest_zone, 0):
                self.highest_zone = active_zone_name
                if frame is not None:
                    self.detection_frame = frame.copy() 
                    # Natychmiast aktualizujemy/zapisujemy klatkę o wyższym priorytecie
                    safe_timestamp = self.start_timestamp_str.replace(":", "-")
                    self.log_queue.put({
                        "type": "save_image",
                        "frame": self.detection_frame,
                        "safe_timestamp": safe_timestamp,
                        "det_id": det_id
                    })
            
            self.session_accuracies.append(current_acc)
            self.session_fps.append(current_fps)
            self.session_inference_times.append(current_inf_time)

        elif not detection_bool and self.is_active:
            duration_s = time.time() - self.start_time
            end_timestamp_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")              
            
            avg_accuracy = sum(self.session_accuracies) / len(self.session_accuracies) if self.session_accuracies else current_acc
            avg_fps = sum(self.session_fps) / len(self.session_fps) if self.session_fps else current_fps
            avg_inference = sum(self.session_inference_times) / len(self.session_inference_times) if self.session_inference_times else current_inf_time
            
            if duration_s < config.MIN_DETECTION_TIME_S:
                status = "requires checking"
            else:
                status = "detection"

            safe_timestamp = self.start_timestamp_str.replace(":", "-")
            
            row = [
                det_id,
                camera_id,
                status,
                self.highest_zone,
                model_name,
                hardware_setup,
                self.start_timestamp_str,
                end_timestamp_str,
                round(duration_s, 2),
                round(avg_fps, 1),
                round(avg_accuracy, 2),
                round(avg_inference, 1),
                robot_x,
                robot_y,
                robot_z
            ]
            
            # Przekazujemy wiersz do zapisu w Excelu (zdjęcie zostało już wysłane wcześniej)
            self.log_queue.put({
                "type": "log",
                "row": row,
                "frame": self.detection_frame,
                "safe_timestamp": safe_timestamp,
                "det_id": det_id
            })
            
            self.next_id = det_id + 1
            self.is_active = False
            self.highest_zone = "NONE"
            self.detection_frame = None 
            self.session_accuracies.clear()
            self.session_fps.clear()
            self.session_inference_times.clear()

    def _apply_formatting(self, ws):
        font_naglowek = Font(name="Calibri", size=11, bold=True)
        wyrownanie_srodek = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.font = font_naglowek
            cell.alignment = wyrownanie_srodek

        if ws.max_row > 1:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(self.headers)):
                for cell in row:
                    cell.alignment = wyrownanie_srodek

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    def _async_writer(self):
        """
        Metoda wątku tła.
        """
        while self.worker_running or not self.log_queue.empty():
            try:
                item = self.log_queue.get(timeout=1.0)
            except queue.Empty:
                continue

            if item is None:
                self.log_queue.task_done()
                break

            # Zapytanie o natychmiastowy zapis obrazu (momentalnie po najechaniu na najwyższą strefę)
            if item["type"] == "save_image":
                frame = item["frame"]
                safe_timestamp = item["safe_timestamp"]
                det_id = item["det_id"]
                try:
                    if frame is not None:
                        base_dir = os.path.dirname(self.log_file_path)
                        folder_name, _ = os.path.splitext(os.path.basename(self.log_file_path))
                        target_folder = os.path.join(base_dir, folder_name)
                        os.makedirs(target_folder, exist_ok=True)
                        img_filename = f"{det_id} {safe_timestamp}.jpg"
                        img_path = os.path.join(target_folder, img_filename)
                        cv2.imwrite(img_path, frame)
                except Exception as e:
                    print(f"[Excel][ERROR] Asynchroniczny błąd zapisu obrazu w locie: {e}")
                self.log_queue.task_done()

            # Zapis rekordu do arkusza Excel
            elif item["type"] == "log":
                row = item["row"]
                frame = item["frame"]
                safe_timestamp = item["safe_timestamp"]
                det_id = item["det_id"]
                cache_path = self.log_file_path + ".cache"
                
                try:
                    if os.path.exists(self.log_file_path):
                        wb = load_workbook(self.log_file_path)
                        ws = wb.active
                    else:
                        wb = Workbook()
                        ws = wb.active
                        ws.title = "datalogs"
                        ws.append(self.headers)
                    
                    ws.append(row)
                    self._apply_formatting(ws)
                    wb.save(self.log_file_path)
                        
                except PermissionError:
                    print(f"\n[Excel][ALERT] Plik {os.path.basename(self.log_file_path)} zablokowany. Zapisuję awaryjnie do cache.")
                    try:
                        existing_buffer = []
                        existing_images = {}
                        if os.path.exists(cache_path):
                            with open(cache_path, 'rb') as f:
                                old_cache = pickle.load(f)
                                existing_buffer = old_cache.get('buffer', [])
                                existing_images = old_cache.get('image_buffer', {})
                        
                        existing_buffer.append(row)
                        if frame is not None:
                            existing_images[det_id] = (frame, safe_timestamp)
                            
                        with open(cache_path, 'wb') as f:
                            pickle.dump({'buffer': existing_buffer, 'image_buffer': existing_images}, f)
                    except Exception as ce:
                        print(f"[Excel][ERROR] Błąd zapisu cache: {ce}")
                except Exception as e:
                    print(f"[Excel][ERROR] Asynchroniczny błąd zapisu wiersza Excel: {e}")
                
                self.log_queue.task_done()

    def save_buffer(self, filename):
        if self.is_active:
            print("[Excel][ZAMYKANIE] Wykryto aktywną detekcję. Wymuszam zapis wiersza...")
            avg_acc = sum(self.session_accuracies) / len(self.session_accuracies) if self.session_accuracies else 0.0
            avg_fps = sum(self.session_fps) / len(self.session_fps) if self.session_fps else 30.0
            avg_inf = sum(self.session_inference_times) / len(self.session_inference_times) if self.session_inference_times else 0.0
            self.acquisition(
                detection_bool=False,
                camera_id=config.CAMERA_ID,
                current_acc=avg_acc,
                model_name=config.MODEL_PATH_NAME,
                hardware_setup=config.HARDWARE_SETUP_STR,
                current_fps=avg_fps,
                current_inf_time=avg_inf,
                robot_x=0.0,
                robot_y=0.0,
                robot_z=0.0,
                active_zone_name=self.highest_zone,
                det_id=self.current_det_id
            )
        self.worker_running = False
        self.log_queue.put(None)
        self.worker_thread.join()